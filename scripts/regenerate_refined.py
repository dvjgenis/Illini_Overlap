#!/usr/bin/env python3
"""Regenerate canonical public Programs_Minors.xlsx from raw data."""
import csv
import re
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAWDATA_DIR = PROJECT_ROOT / "rawdata"
DST = Path(__file__).resolve().parents[1] / "public" / "Programs_Minors.xlsx"

NEW_FIELDS = [
    "program_type",
    "name",
    "code",
    "college",
    "department",
    "credit_hours",
    "required_courses",
    "elective_courses",
    "electives_required",
    "prerequisites",
    "prerequisites_courses",
    "advisory_approval_required",
    "up_to_date",
    "effective_catalog_term",
    "effective_catalog",
    "overlap_info",
    "excluded_majors",
    "requirements_text",
    "notes",
    "status",
    "description",
]


def split_course_list(value: str) -> list[str]:
    if not value:
        return []
    return [x.strip() for x in re.split(r"[|;,\n]+", value) if x.strip()]


def normalize_course_code(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", value or "").strip().upper()
    m = re.match(r"^([A-Z]{2,5})\s*[-_]?\s*(\d{3,4}[A-Z]?)$", cleaned)
    if not m:
        return ""
    return f"{m.group(1)} {m.group(2)}"


def extract_course_codes_from_list(value: str) -> list[str]:
    tokens = split_course_list(value)
    out: list[str] = []
    seen: set[str] = set()
    last_dept = ""

    def push(code: str) -> None:
        norm = normalize_course_code(code)
        if not norm or norm in seen:
            return
        seen.add(norm)
        out.append(norm)

    for token in tokens:
        cleaned = normalize_text(token)
        if not cleaned:
            continue

        explicit = list(
            re.finditer(
                r"([A-Z]{2,5})\s*[-_]?\s*(\d{3,4}[A-Z]?)(?:\s*/\s*(\d{3,4}[A-Z]?))?",
                cleaned.upper(),
            )
        )
        if explicit:
            for m in explicit:
                dept = m.group(1)
                first = m.group(2)
                second = m.group(3)
                push(f"{dept} {first}")
                if second:
                    push(f"{dept} {second}")
                last_dept = dept
            continue

        if re.fullmatch(r"\d{3,4}[A-Z]?", cleaned.upper()) and last_dept:
            push(f"{last_dept} {cleaned.upper()}")

    return out


def join_pipe(values: list[str]) -> str:
    return "|".join(v for v in values if v)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\u00A0", " ")
        .replace("\uFFFD", " ")
        .replace("�", " ")
        .replace("\r", " ")
        .replace("\n", " ")
        .strip()
    )


def is_sentinel_none(value: str) -> bool:
    """Return True for values that are effectively empty/placeholder."""
    return value.strip().lower() in ("none", "n/a", "na", "")


def estimate_total_courses(credit_hours: str) -> int | None:
    if not credit_hours:
        return None
    m = re.search(r"(\d+)(?:\s*-\s*(\d+))?", credit_hours)
    if not m:
        return None
    try:
        min_hours = int(m.group(1))
    except ValueError:
        return None
    return max(1, (min_hours + 2) // 3)


def parse_electives_required(value: str) -> int | None:
    if not value:
        return None
    m = re.search(r"\d+", value)
    if not m:
        return None
    return int(m.group(0))


def normalize_advisory(raw: str) -> str:
    """Normalize advisory_approval_required to Yes/No."""
    low = raw.strip().lower()
    if low.startswith("yes"):
        return "Yes"
    return "No"


def pick_latest_raw() -> Path:
    """Pick the most recently modified raw file (Excel or CSV)."""
    candidates = []
    for ext in ("*.xlsx", "*.xls", "*.csv"):
        candidates.extend(RAWDATA_DIR.glob(ext))
    candidates = [p for p in candidates if p.is_file()]
    if not candidates:
        raise SystemExit(f"No Excel or CSV files found in rawdata directory: {RAWDATA_DIR}")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def read_raw_excel(path: Path) -> list[dict[str, str]]:
    """Read raw Excel file into list of dicts."""
    if not HAS_OPENPYXL:
        raise SystemExit("openpyxl is required for Excel support. Install with: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    result = []
    for row in rows[1:]:
        values = list(row) if row else []
        while len(values) < len(headers):
            values.append(None)
        values = values[: len(headers)]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        result.append(dict(zip(headers, [normalize_text(v) for v in values])))
    return result


def read_raw_csv(path: Path) -> list[dict[str, str]]:
    """Read raw CSV file into list of dicts."""
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        return [{k: normalize_text(v) for k, v in row.items()} for row in reader]


def main() -> None:
    raw_path = pick_latest_raw()
    suffix = raw_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        rows = read_raw_excel(raw_path)
    else:
        rows = read_raw_csv(raw_path)

    seen_keys: set[tuple[str, str]] = set()
    out_rows: list[dict[str, str]] = []

    for row in rows:
        name = row.get("name", "").strip()
        if not name:
            continue

        program_type = (row.get("program_type", "") or "").strip().lower() or "other"
        dedup_key = (name.lower(), program_type)
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)

        req_list = extract_course_codes_from_list(row.get("required_courses", ""))
        elec_list = extract_course_codes_from_list(row.get("elective_courses", ""))

        credit_hours = row.get("credit_hours", "")
        total_est = estimate_total_courses(credit_hours)

        electives_required = parse_electives_required(row.get("electives_required", ""))
        if electives_required is None:
            if elec_list:
                if total_est is not None:
                    needed = max(1, total_est - len(req_list))
                    electives_required = min(3, needed, len(elec_list))
                else:
                    electives_required = min(2, len(elec_list))
            elif req_list:
                electives_required = 0

        excluded = row.get("excluded_majors", "")
        if is_sentinel_none(excluded):
            excluded = ""

        advisory = normalize_advisory(row.get("advisory_approval_required", ""))

        raw_req_text = row.get("requirements_text", "")
        extra_parts: list[str] = []
        for key in ("prerequisites", "overlap_info", "notes", "description"):
            val = (row.get(key, "") or "").strip()
            if val and not is_sentinel_none(val):
                extra_parts.append(val)
        if raw_req_text and not is_sentinel_none(raw_req_text):
            requirements_text = raw_req_text
            if extra_parts:
                requirements_text += " | " + " | ".join(extra_parts)
        else:
            requirements_text = " | ".join(extra_parts)

        out_row = {k: row.get(k, "") for k in NEW_FIELDS if k in row}
        for k in NEW_FIELDS:
            if k not in out_row:
                out_row[k] = ""

        out_row["required_courses"] = join_pipe(req_list)
        out_row["elective_courses"] = join_pipe(elec_list)
        out_row["electives_required"] = "" if electives_required is None else str(electives_required)
        out_row["excluded_majors"] = excluded
        out_row["requirements_text"] = requirements_text
        out_row["advisory_approval_required"] = advisory

        out_rows.append(out_row)

    # Write Excel
    if not HAS_OPENPYXL:
        raise SystemExit("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Programs"
    ws.append(NEW_FIELDS)
    for row in out_rows:
        ws.append([row.get(k, "") for k in NEW_FIELDS])
    wb.save(DST)

    print(f"Regenerated from raw data ({raw_path.name}): {DST}")


if __name__ == "__main__":
    main()
